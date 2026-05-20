import csv
import re
from collections import defaultdict, OrderedDict
from pathlib import Path

from PySide6.QtCore import Qt
from PySide6.QtGui import QColor, QFont, QBrush, QAction
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QFrame, QHeaderView,
    QFileDialog, QMessageBox, QDialog, QMenu, QInputDialog, QDoubleSpinBox,
)

from src.constants import AIR_IMPORT_BUCKETS, FCL_BUCKETS, LCL_BUCKETS
from src.pages.comparison_helpers import (
    _FLAT_UNITS, _UNIT_ABBREVS, _UNIT_FROM_ABBREV,
    _is_flat, _unit_abbrev, _fmt_total_usd, _fmt_total_inr,
    _cell_display, _parse_cell_input,
)
from src.pages.comparison_dialogs import (
    _AddChargeDialog, RateFetchWorker, CustomRateDialog, _POPUP_STYLE,
)


# ---------------------------------------------------------------------------
# Comparison Page
# ---------------------------------------------------------------------------
_NON_EDIT = Qt.ItemIsEnabled
_EDITABLE = Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable

_HDR_BG   = "#E8EEF6"
_HDR_FG   = "#1565C0"
_TOT_BG   = "#F0F4F8"
_GREEN    = "#C8E6C9"
_GREEN_FG = "#1B5E20"
_YELLOW   = "#FFF8E1"
_DARK     = "#1E2A3A"
_DASH_FG  = "#90A4AE"
_WHITE    = "#FFFFFF"
_VHDR_BG  = "#2C3E50"
_VHDR_FG  = "#ECEFF1"
_INFO_BG  = "#EEF2F7"

_POPUP_STYLE = """
    QDialog {
        background: #F8FAFC;
        color: #1E2A3A;
    }
    QDialog QLabel {
        color: #1E2A3A;
        font-size: 12px;
    }
    QDialog QLineEdit,
    QDialog QComboBox,
    QDialog QDoubleSpinBox {
        min-height: 28px;
        padding: 0 8px;
        color: #1E2A3A;
        background: #FFFFFF;
        border: 1px solid #B0BEC5;
        border-radius: 4px;
        selection-background-color: #BBDEFB;
        selection-color: #1E2A3A;
    }
    QDialog QLineEdit:focus,
    QDialog QComboBox:focus,
    QDialog QDoubleSpinBox:focus {
        border: 1px solid #1976D2;
    }
    QDialog QComboBox::drop-down,
    QDialog QDoubleSpinBox::up-button,
    QDialog QDoubleSpinBox::down-button {
        border: none;
        width: 18px;
    }
    QDialog QComboBox QAbstractItemView {
        color: #1E2A3A;
        background: #FFFFFF;
        border: 1px solid #CFD8DC;
        selection-background-color: #E3F2FD;
        selection-color: #1E2A3A;
    }
    QDialog QPushButton {
        background: #FFFFFF;
        color: #1E2A3A;
        border: 1px solid #B0BEC5;
        border-radius: 5px;
        padding: 6px 14px;
        font-size: 12px;
    }
    QDialog QPushButton:hover {
        background: #EEF2F7;
    }
    QDialog QPushButton:disabled {
        background: #E9EEF3;
        color: #90A4AE;
        border-color: #CFD8DC;
    }
    QMenu {
        background: #FFFFFF;
        color: #1E2A3A;
        border: 1px solid #CFD8DC;
        padding: 4px;
    }
    QMenu::item {
        padding: 6px 18px;
    }
    QMenu::item:selected {
        background: #E3F2FD;
        color: #1E2A3A;
    }
"""

_SPIN_STYLE = (
    "QDoubleSpinBox{border:1px solid #B0BEC5;border-radius:4px;"
    "padding:0 4px;font-size:12px;color:#1E2A3A;background:#FFFFFF;}"
)

_MODE_BTN_BASE = (
    "QPushButton{border:1px solid #B0BEC5;border-radius:5px;"
    "padding:0 14px;font-size:12px;font-weight:600;}"
)
_MODE_COLORS = {
    "air": ("#1976D2", "#1565C0"),
    "fcl": ("#00796B", "#00695C"),
    "lcl": ("#2E7D32", "#1B5E20"),
}


class ComparisonPage(QWidget):
    def __init__(self, app_window, parent=None):
        super().__init__(parent)
        self.app = app_window
        # Last table vendors (column -> VendorData)
        self._table_vendors: list = []
        self._worker: RateFetchWorker | None = None
        self._popup_on_done = False
        self._row_specs: list = []
        self._charge_weight: float = 0.0
        self._current_mode: str = "air"
        self._volume_cbm: float = 0.0
        self._gross_kg: float = 0.0
        self._chargeable_cbm: float = 0.0
        self._custom_rates: dict = {}
        self._unit_quantities: dict[str, float] = {}
        self._unit_spins: dict[str, QDoubleSpinBox] = {}
        self._build_ui()

    # ------------------------------------------------------------------
    def _build_ui(self):
        self.setStyleSheet("""
            QToolTip {
                background-color: #1E2A3A;
                color: #FFFFFF;
                border: 1px solid #354A5E;
                padding: 5px 10px;
                font-size: 12px;
                border-radius: 4px;
            }
        """ + _POPUP_STYLE)

        root = QVBoxLayout(self)
        root.setContentsMargins(32, 28, 32, 20)
        root.setSpacing(14)

        # Title row
        title_row = QHBoxLayout()
        title = QLabel("Vendor Comparison")
        title.setStyleSheet("font-size:20px; font-weight:700; color:#1E2A3A;")
        title_row.addWidget(title)
        title_row.addStretch()

        self.rate_label = QLabel("Fetching exchange rates…")
        self.rate_label.setStyleSheet("font-size:12px; color:#607080;")
        title_row.addWidget(self.rate_label)

        refresh_btn = QPushButton("Refresh Rates")
        refresh_btn.setFixedHeight(28)
        refresh_btn.setStyleSheet(
            "QPushButton{background:#EEF2F7;color:#1E2A3A;border:1px solid #B0BEC5;"
            "border-radius:5px;font-size:11px;padding:0 10px;}"
            "QPushButton:hover{background:#E3F2FD;}"
        )
        refresh_btn.clicked.connect(self._fetch_rates)
        title_row.addWidget(refresh_btn)

        self._custom_rate_btn = QPushButton("Custom Rates")
        self._custom_rate_btn.setFixedHeight(28)
        self._custom_rate_btn.setStyleSheet(
            "QPushButton{background:#EEF2F7;color:#1E2A3A;border:1px solid #B0BEC5;"
            "border-radius:5px;font-size:11px;padding:0 10px;}"
            "QPushButton:hover{background:#E3F2FD;}"
        )
        self._custom_rate_btn.clicked.connect(self._open_custom_rates)
        title_row.addWidget(self._custom_rate_btn)
        root.addLayout(title_row)

        # Mode bar (shown when more than one mode has data)
        self._mode_bar = QFrame()
        self._mode_bar.setVisible(False)
        mb_layout = QHBoxLayout(self._mode_bar)
        mb_layout.setContentsMargins(0, 0, 0, 0)
        mb_layout.setSpacing(4)

        self._mode_btns: dict[str, QPushButton] = {}
        for mode, label in [("air", "Air Freight"), ("fcl", "FCL (Ocean)"), ("lcl", "LCL (Ocean)")]:
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setFixedHeight(30)
            btn.clicked.connect(lambda _checked, m=mode: self._set_mode(m))
            self._mode_btns[mode] = btn
            mb_layout.addWidget(btn)
        mb_layout.addStretch()
        root.addWidget(self._mode_bar)
        self._apply_mode_btn_styles()

        sub = QLabel(
            "Charges shown in original currency. "
            "Totals split by flat charges + per-unit rates. "
            "Double-click a charge cell to edit. Hover yellow cells for remarks."
        )
        sub.setStyleSheet("font-size:13px; color:#607080;")
        root.addWidget(sub)

        # Toolbar
        toolbar = QHBoxLayout()

        add_vendor_btn = QPushButton("Add Vendor")
        add_vendor_btn.setFixedHeight(28)
        add_vendor_btn.setStyleSheet(self._sm_btn_style())
        add_vendor_btn.clicked.connect(self._add_vendor_column)
        toolbar.addWidget(add_vendor_btn)

        add_charge_btn = QPushButton("Add Charge")
        add_charge_btn.setFixedHeight(28)
        add_charge_btn.setStyleSheet(self._sm_btn_style())
        add_charge_btn.clicked.connect(self._add_charge_row)
        toolbar.addWidget(add_charge_btn)

        preview_btn = QPushButton("Preview Imported Mails")
        preview_btn.setFixedHeight(28)
        preview_btn.setStyleSheet(self._sm_btn_style())
        preview_btn.clicked.connect(self._preview_imported)
        toolbar.addWidget(preview_btn)

        toolbar.addStretch()

        # Air inputs: chargeable weight
        self._air_inputs = QFrame()
        ai_layout = QHBoxLayout(self._air_inputs)
        ai_layout.setContentsMargins(0, 0, 0, 0)
        ai_layout.setSpacing(6)
        wt_label = QLabel("Chargeable Weight:")
        wt_label.setStyleSheet("font-size:12px; color:#607080;")
        ai_layout.addWidget(wt_label)
        self._weight_spin = QDoubleSpinBox()
        self._weight_spin.setRange(0, 99999)
        self._weight_spin.setDecimals(1)
        self._weight_spin.setSuffix(" KG")
        self._weight_spin.setSpecialValueText("—")
        self._weight_spin.setFixedWidth(120)
        self._weight_spin.setFixedHeight(28)
        self._weight_spin.setStyleSheet(_SPIN_STYLE)
        self._weight_spin.valueChanged.connect(self._on_weight_changed)
        ai_layout.addWidget(self._weight_spin)
        toolbar.addWidget(self._air_inputs)

        # LCL inputs: volume CBM + weight KG → auto chargeable CBM
        self._lcl_inputs = QFrame()
        self._lcl_inputs.setVisible(False)
        lcl_layout = QHBoxLayout(self._lcl_inputs)
        lcl_layout.setContentsMargins(0, 0, 0, 0)
        lcl_layout.setSpacing(6)

        cbm_label = QLabel("Volume:")
        cbm_label.setStyleSheet("font-size:12px; color:#607080;")
        lcl_layout.addWidget(cbm_label)
        self._cbm_spin = QDoubleSpinBox()
        self._cbm_spin.setRange(0, 9999)
        self._cbm_spin.setDecimals(3)
        self._cbm_spin.setSuffix(" CBM")
        self._cbm_spin.setSpecialValueText("—")
        self._cbm_spin.setFixedWidth(130)
        self._cbm_spin.setFixedHeight(28)
        self._cbm_spin.setStyleSheet(_SPIN_STYLE)
        self._cbm_spin.valueChanged.connect(self._on_lcl_inputs_changed)
        lcl_layout.addWidget(self._cbm_spin)

        lcl_layout.addSpacing(8)
        kg_lbl = QLabel("Weight:")
        kg_lbl.setStyleSheet("font-size:12px; color:#607080;")
        lcl_layout.addWidget(kg_lbl)
        self._lcl_kg_spin = QDoubleSpinBox()
        self._lcl_kg_spin.setRange(0, 99999)
        self._lcl_kg_spin.setDecimals(1)
        self._lcl_kg_spin.setSuffix(" KG")
        self._lcl_kg_spin.setSpecialValueText("—")
        self._lcl_kg_spin.setFixedWidth(120)
        self._lcl_kg_spin.setFixedHeight(28)
        self._lcl_kg_spin.setStyleSheet(_SPIN_STYLE)
        self._lcl_kg_spin.valueChanged.connect(self._on_lcl_inputs_changed)
        lcl_layout.addWidget(self._lcl_kg_spin)

        lcl_layout.addSpacing(10)
        self._chargeable_lbl = QLabel("Chargeable: —")
        self._chargeable_lbl.setStyleSheet(
            "font-size:12px; color:#1E2A3A; font-weight:600;"
        )
        lcl_layout.addWidget(self._chargeable_lbl)
        toolbar.addWidget(self._lcl_inputs)

        # Dynamic per-unit quantity inputs (Per AWB, Per BL, etc.)
        self._extra_inputs = QFrame()
        self._extra_inputs.setVisible(False)
        ei_layout = QHBoxLayout(self._extra_inputs)
        ei_layout.setContentsMargins(0, 0, 0, 0)
        ei_layout.setSpacing(4)
        toolbar.addWidget(self._extra_inputs)

        root.addLayout(toolbar)

        # Table
        self.table = QTableWidget()
        self.table.setEditTriggers(
            QTableWidget.DoubleClicked | QTableWidget.AnyKeyPressed
        )
        self.table.setSelectionBehavior(QTableWidget.SelectItems)
        self.table.setAlternatingRowColors(False)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.verticalHeader().setVisible(False)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._on_context_menu)
        self.table.itemChanged.connect(self._on_cell_changed)
        self.table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #DCE3EA;
                border-radius: 6px;
                gridline-color: #E8EDF2;
                font-size: 13px;
                background: #FFFFFF;
                color: #1E2A3A;
            }
            QTableWidget::item { padding: 6px 12px; color: #1E2A3A; }
            QHeaderView::section {
                background: #1E2A3A;
                color: #FFFFFF;
                font-size: 11px;
                font-weight: 700;
                padding: 7px 10px;
                border: none;
                border-right: 1px solid #354A5E;
            }
            QTableWidget::item:selected { background: #E3F2FD; color: #1E2A3A; }
            QTableWidget QLineEdit {
                background: #FFFFFF;
                color: #1E2A3A;
                border: 1px solid #90CAF9;
                border-radius: 3px;
                padding: 4px 8px;
                selection-background-color: #BBDEFB;
                selection-color: #1E2A3A;
            }
            QTableWidget QLineEdit:focus {
                background: #FFFFFF;
                color: #1E2A3A;
            }
        """)
        root.addWidget(self.table, 1)

        # Legend
        legend_row = QHBoxLayout()
        green_box = QLabel()
        green_box.setFixedSize(14, 14)
        green_box.setStyleSheet(
            "background:#C8E6C9;border:1px solid #A5D6A7;border-radius:2px;"
        )
        legend_row.addWidget(green_box)
        lbl_green = QLabel("Lowest cost in row")
        lbl_green.setStyleSheet("font-size:12px; font-weight:600; color:#1E2A3A;")
        legend_row.addWidget(lbl_green)
        legend_row.addSpacing(20)
        yellow_box = QLabel()
        yellow_box.setFixedSize(14, 14)
        yellow_box.setStyleSheet(
            "background:#FFF8E1;border:1px solid #FFE082;border-radius:2px;"
        )
        legend_row.addWidget(yellow_box)
        lbl_yellow = QLabel("Has remarks — hover to view")
        lbl_yellow.setStyleSheet("font-size:12px; font-weight:600; color:#1E2A3A;")
        legend_row.addWidget(lbl_yellow)
        legend_row.addStretch()
        root.addLayout(legend_row)

        # Bottom nav
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("color:#DCE3EA;")
        root.addWidget(sep)

        nav = QHBoxLayout()
        back_btn = QPushButton("← Back to Mapping")
        back_btn.setFixedHeight(36)
        back_btn.setStyleSheet(self._sec_btn_style())
        back_btn.clicked.connect(lambda: self.app.go_to_mapping())
        nav.addWidget(back_btn)
        # Start New Comparison button — clears current session and returns to Import
        start_new_btn = QPushButton("Start New Comparison")
        start_new_btn.setFixedHeight(36)
        start_new_btn.setStyleSheet(self._sec_btn_style())
        start_new_btn.clicked.connect(self._start_new_comparison)
        nav.addWidget(start_new_btn)
        nav.addStretch()

        export_csv_btn = QPushButton("Export CSV")
        export_csv_btn.setFixedHeight(36)
        export_csv_btn.setStyleSheet(self._sec_btn_style())
        export_csv_btn.clicked.connect(self._export_csv)
        nav.addWidget(export_csv_btn)

        export_xl_btn = QPushButton("Export Excel")
        export_xl_btn.setFixedHeight(36)
        export_xl_btn.setStyleSheet(self._pri_btn_style())
        export_xl_btn.clicked.connect(self._export_excel)
        nav.addWidget(export_xl_btn)
        root.addLayout(nav)

    def _start_new_comparison(self):
        resp = QMessageBox.question(
            self, "Start New Comparison",
            "This will clear the current comparison and return you to the Import page. Continue?",
            QMessageBox.Yes | QMessageBox.No
        )
        if resp != QMessageBox.Yes:
            return

        # Clear in-memory vendors
        self.app.vendors.clear()

        # Clear import page file rows (no extra confirmation)
        ip = getattr(self.app, "import_page", None)
        if ip:
            for p in list(ip._file_rows.keys()):
                ip._remove_file(p)

        # Navigate back to import
        self.app.go_to_import()

    def _preview_imported(self):
        try:
            from src.pages.preview_mails import PreviewMailsDialog
        except Exception:
            QMessageBox.warning(self, "Unable to preview", "Preview component unavailable.")
            return
        dlg = PreviewMailsDialog(self.app, self)
        dlg.exec()

    # ------------------------------------------------------------------
    # Mode bar helpers
    # ------------------------------------------------------------------
    def _apply_mode_btn_styles(self):
        for mode, btn in self._mode_btns.items():
            active = btn.isChecked()
            if active:
                bg, hover = _MODE_COLORS[mode]
                btn.setStyleSheet(
                    f"QPushButton{{background:{bg};color:white;border:none;"
                    f"border-radius:5px;padding:0 14px;font-size:12px;font-weight:600;}}"
                    f"QPushButton:hover{{background:{hover};}}"
                )
            else:
                btn.setStyleSheet(
                    "QPushButton{background:#EEF2F7;color:#607080;"
                    "border:1px solid #B0BEC5;border-radius:5px;"
                    "padding:0 14px;font-size:12px;font-weight:600;}"
                    "QPushButton:hover{background:#E3F2FD;color:#1E2A3A;}"
                )

    def _update_mode_tabs(self):
        all_vds = list(self.app.vendors.values())
        air_has = any(vd.quote_type == "air" for vd in all_vds)
        fcl_has = any(vd.quote_type == "fcl" for vd in all_vds)
        lcl_has = any(vd.quote_type == "lcl" for vd in all_vds)

        available = [m for m, has in [("air", air_has), ("fcl", fcl_has), ("lcl", lcl_has)] if has]

        if len(available) > 1:
            self._mode_bar.setVisible(True)
        else:
            self._mode_bar.setVisible(False)

        if available and self._current_mode not in available:
            self._current_mode = available[0]

        for mode, btn in self._mode_btns.items():
            btn.setChecked(mode == self._current_mode)
        self._apply_mode_btn_styles()

    def _set_mode(self, mode: str):
        self._current_mode = mode
        for m, btn in self._mode_btns.items():
            btn.setChecked(m == mode)
        self._apply_mode_btn_styles()
        self._air_inputs.setVisible(mode == "air")
        self._lcl_inputs.setVisible(mode == "lcl")
        self._build_table()

    # ------------------------------------------------------------------
    def load(self):
        self._update_mode_tabs()
        self._air_inputs.setVisible(self._current_mode == "air")
        self._lcl_inputs.setVisible(self._current_mode == "lcl")
        # Pre-populate weight from import page if user hasn't overridden it here
        w = getattr(self.app, "chargeable_weight", 0.0)
        if w > 0 and self._weight_spin.value() == 0:
            self._weight_spin.setValue(w)
            self._charge_weight = w
        svc = self.app.currency_service
        if svc.last_updated is None:
            # No rates in memory yet — fetch (cache first, live API as fallback)
            self._fetch_rates(show_popup=True)
        else:
            # Rates ready — rebuild table so any mapping changes are reflected
            self.rate_label.setText(svc.rate_display())
            self.rate_label.setStyleSheet("font-size:12px;color:#607080;")
            self._apply_custom_rate_style()
            self._build_table()

    def refresh_service_state(self):
        if self._custom_rates:
            self._apply_custom_rate_style()
            return
        self.rate_label.setText(self.app.currency_service.rate_display())
        self.rate_label.setStyleSheet("font-size:12px;color:#607080;")


    def _fetch_rates(self, show_popup: bool = False):
        """Explicitly re-fetch rates (clears custom overrides). Used on first load and Refresh."""
        self._custom_rates = {}
        self._apply_custom_rate_style()
        self._popup_on_done = show_popup
        self.rate_label.setText("Fetching exchange rates…")
        self._worker = RateFetchWorker(self.app.currency_service, self)
        self._worker.done.connect(self._on_rates_fetched)
        self._worker.start()

    def _on_rates_fetched(self, _ok: bool):
        self.rate_label.setText(self.app.currency_service.rate_display())
        self.rate_label.setStyleSheet("font-size:12px;color:#607080;")
        self._popup_on_done = False
        self._apply_custom_rate_style()
        self._build_table()

    def _on_weight_changed(self, value: float):
        self._charge_weight = value
        self._update_totals()

    def _on_lcl_inputs_changed(self):
        cbm = self._cbm_spin.value()
        kg = self._lcl_kg_spin.value()
        chargeable = max(cbm, kg / 1000.0) if (cbm > 0 or kg > 0) else 0.0
        self._volume_cbm = cbm
        self._gross_kg = kg
        self._chargeable_cbm = chargeable
        if chargeable > 0:
            self._chargeable_lbl.setText(f"Chargeable: {chargeable:.3f} CBM")
        else:
            self._chargeable_lbl.setText("Chargeable: —")
        self._update_totals()

    # ------------------------------------------------------------------
    # Custom rate helpers
    # ------------------------------------------------------------------
    def _to_usd(self, amount: float, currency: str) -> float:
        if self._custom_rates:
            rate = self._custom_rates.get(currency.upper())
            if rate and rate > 0:
                return amount / rate
        return self.app.currency_service.to_usd(amount, currency)

    def _get_usd_to_inr(self) -> float:
        if self._custom_rates:
            rate = self._custom_rates.get("INR")
            if rate and rate > 0:
                return rate
        return self.app.currency_service.usd_to_inr

    def _open_custom_rates(self):
        currencies: set = set()
        for vd in self._table_vendors:
            for ch in vd.charges:
                if ch.currency and ch.currency.upper() != "USD":
                    currencies.add(ch.currency.upper())
        currencies.add("INR")

        svc = self.app.currency_service
        live_rates = {c: svc.rates.get(c, 1.0) for c in currencies}

        dlg = CustomRateDialog(
            live_rates=live_rates,
            currencies=sorted(currencies),
            current_custom=self._custom_rates,
            parent=self,
        )
        if dlg.exec() != QDialog.Accepted:
            return

        self._custom_rates = dlg.get_custom_rates()
        self._apply_custom_rate_style()
        self._build_table()

    def _apply_custom_rate_style(self):
        if self._custom_rates:
            self.rate_label.setText("Custom rates active")
            self.rate_label.setStyleSheet(
                "font-size:12px;color:#E65100;font-weight:600;"
                "background:#FFF3E0;border:1px solid #FFB74D;"
                "border-radius:4px;padding:2px 8px;"
            )
            self._custom_rate_btn.setStyleSheet(
                "QPushButton{background:#E65100;color:white;border:none;"
                "border-radius:5px;font-size:11px;font-weight:600;padding:0 10px;}"
                "QPushButton:hover{background:#BF360C;}"
            )
            self._custom_rate_btn.setText("Custom Rates ✓")
        else:
            self.rate_label.setText(self.app.currency_service.rate_display())
            self.rate_label.setStyleSheet("font-size:12px;color:#607080;")
            self._custom_rate_btn.setStyleSheet(
                "QPushButton{background:#EEF2F7;color:#1E2A3A;border:1px solid #B0BEC5;"
                "border-radius:5px;font-size:11px;padding:0 10px;}"
                "QPushButton:hover{background:#E3F2FD;}"
            )
            self._custom_rate_btn.setText("Custom Rates")

    # ------------------------------------------------------------------
    # Dynamic unit-quantity inputs
    # ------------------------------------------------------------------
    def _refresh_unit_inputs(self):
        mode = self._current_mode
        already_handled = set(_FLAT_UNITS)
        if mode == "air":
            already_handled.add("per kg")
        elif mode == "lcl":
            already_handled.update({"per cbm", "per ton", "per kg"})

        needed: dict[str, str] = {}
        for vd in self._table_vendors:
            for ch in vd.charges:
                if ch.if_applicable:
                    continue
                u_lower = (ch.unit_of_measurement or "").lower().strip()
                if u_lower and u_lower not in already_handled and u_lower not in needed:
                    needed[u_lower] = ch.unit_of_measurement.strip()

        layout = self._extra_inputs.layout()
        for i in reversed(range(layout.count())):
            item = layout.takeAt(i)
            if item and item.widget():
                item.widget().deleteLater()
        self._unit_spins.clear()

        if not needed:
            self._extra_inputs.setVisible(False)
            return

        for u_lower, u_display in sorted(needed.items()):
            abbrev = _unit_abbrev(u_display)
            lbl = QLabel(f"{abbrev}:")
            lbl.setStyleSheet("font-size:12px; color:#607080;")
            layout.addWidget(lbl)
            spin = QDoubleSpinBox()
            spin.setRange(0, 9999)
            spin.setDecimals(0)
            spin.setValue(self._unit_quantities.get(u_lower, 1.0))
            spin.setSuffix(f" {abbrev}")
            spin.setFixedWidth(90)
            spin.setFixedHeight(28)
            spin.setStyleSheet(_SPIN_STYLE)
            spin.valueChanged.connect(lambda v, u=u_lower: self._on_unit_qty_changed(u, v))
            layout.addWidget(spin)
            layout.addSpacing(8)
            self._unit_spins[u_lower] = spin
            if u_lower not in self._unit_quantities:
                self._unit_quantities[u_lower] = 1.0

        self._extra_inputs.setVisible(True)

    def _on_unit_qty_changed(self, unit: str, value: float):
        self._unit_quantities[unit] = value
        self._update_totals()

    def _needs_actual_total(self) -> bool:
        mode = self._current_mode
        return (
            (mode == "air" and self._charge_weight > 0)
            or (mode == "lcl" and self._chargeable_cbm > 0)
            or bool(self._unit_spins)
        )

    def _compute_actual(self, flat: float, variable: dict[str, float]) -> float:
        mode = self._current_mode
        total = flat
        for unit, rate in variable.items():
            u = unit.lower().strip()
            if mode == "air" and u == "per kg":
                qty = self._charge_weight
            elif mode == "lcl" and u in ("per cbm", "per ton"):
                qty = self._chargeable_cbm
            else:
                qty = self._unit_quantities.get(u, 1.0)
            total += rate * qty
        return total

    # ------------------------------------------------------------------
    # Table building
    # ------------------------------------------------------------------
    _SLAB_RE = re.compile(r'air\s+freight\s*[+\-<]\s*\d+', re.IGNORECASE)

    def _has_applicable_af(self, vd) -> bool:
        """
        Return False when a slab-based vendor has no visible air-freight rate row.

        Auto-marking hides slab rows that do not match the chargeable weight by
        setting if_applicable=True, and renames the matching slab to
        "Air Freight". If the user later goes back to Mapping and manually
        unchecks Optional on a slab row, that explicit override should make the
        quote appear again in Comparison even if the row keeps its slab name.
        """
        af_charges = [ch for ch in vd.charges if ch.category == "AF (Air Freight)"]
        if not af_charges:
            return True  # no AF charges at all — keep

        if getattr(self.app, "chargeable_weight", 0.0) <= 0:
            return True  # weight not set — auto-marking never ran, keep all

        has_slabs = any(self._SLAB_RE.search(ch.name_of_charge) for ch in af_charges)
        if not has_slabs:
            return True  # flat-rate vendor — always keep

        # Slab vendor: keep it if there is any non-optional air-freight rate row,
        # whether it is the auto-normalised "Air Freight" row or a user-restored slab.
        return any(
            not ch.if_applicable and (
                ch.name_of_charge.strip().lower() == "air freight"
                or self._SLAB_RE.search(ch.name_of_charge)
            )
            for ch in af_charges
        )

    def _build_table(self):
        if self._current_mode == "fcl":
            vendors = [vd for vd in self.app.vendors.values() if vd.quote_type == "fcl"]
            self._build_fcl_table(vendors, FCL_BUCKETS)
        elif self._current_mode == "lcl":
            vendors = [vd for vd in self.app.vendors.values() if vd.quote_type == "lcl"]
            self._build_standard_table(vendors, LCL_BUCKETS)
        else:
            vendors = [vd for vd in self.app.vendors.values() if vd.quote_type == "air"]
            vendors = [vd for vd in vendors if self._has_applicable_af(vd)]
            self._build_standard_table(vendors, AIR_IMPORT_BUCKETS)

    def _build_standard_table(self, vendors: list, buckets: list):
        if not vendors:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            return

        bucket_charge_rows: dict[str, list[tuple[str, str]]] = {b: [] for b in buckets}
        seen_per_bucket: dict[str, set] = defaultdict(set)

        for vd in vendors:
            for charge in vd.charges:
                if charge.if_applicable:
                    continue
                bkt = charge.category
                if bkt not in bucket_charge_rows:
                    continue
                name_lower = charge.name_of_charge.strip().lower()
                if name_lower and name_lower not in seen_per_bucket[bkt]:
                    seen_per_bucket[bkt].add(name_lower)
                    bucket_charge_rows[bkt].append(
                        (name_lower, charge.name_of_charge.strip())
                    )

        row_specs: list = []
        if any(vd.airline for vd in vendors):
            row_specs.append(("airline_row",))
        if any(vd.etd for vd in vendors):
            row_specs.append(("info", "etd", "ETD"))
        if any(vd.transit_days for vd in vendors):
            row_specs.append(("info", "transit_days", "Transit Time"))
        if any(vd.free_days_origin > 0 for vd in vendors):
            row_specs.append(("info", "free_days_origin", "Free Days (Origin)"))
        if any(vd.free_days_destination > 0 for vd in vendors):
            row_specs.append(("info", "free_days_destination", "Free Days (Destination)"))
        for bkt in buckets:
            if bucket_charge_rows[bkt]:
                row_specs.append(("header", bkt))
                for name_lower, display in bucket_charge_rows[bkt]:
                    row_specs.append(("charge", bkt, name_lower, display))
        row_specs.append(("total_usd",))
        row_specs.append(("total_inr",))
        self._row_specs = row_specs

        headers = ["Charge"] + [vd.vendor_name for vd in vendors]
        # Track which VendorData maps to each vendor column (used for exports)
        self._table_vendors = list(vendors)
        self._refresh_unit_inputs()
        n_cols = len(headers)

        # column-indexed totals
        flat_totals: dict[int, float] = {c: 0.0 for c in range(1, n_cols)}
        variable_totals: dict[int, dict[str, float]] = {c: {} for c in range(1, n_cols)}

        self.table.blockSignals(True)
        self.table.setColumnCount(n_cols)
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(row_specs))

        bold_font = QFont(); bold_font.setBold(True)
        hdr_font  = QFont(); hdr_font.setBold(True); hdr_font.setPointSize(10)

        for r_idx, spec in enumerate(row_specs):

            if spec[0] == "airline_row":
                self.table.setRowHeight(r_idx, 28)
                lbl = QTableWidgetItem("  Airline")
                lbl.setBackground(QBrush(QColor(_VHDR_BG)))
                lbl.setForeground(QBrush(QColor(_VHDR_FG)))
                lbl.setFont(hdr_font)
                lbl.setFlags(_NON_EDIT)
                self.table.setItem(r_idx, 0, lbl)
                for c_idx, vd in enumerate(vendors, start=1):
                    cell = QTableWidgetItem(vd.airline or "—")
                    cell.setBackground(QBrush(QColor(_VHDR_BG)))
                    cell.setForeground(QBrush(QColor(_VHDR_FG)))
                    cell.setFont(bold_font)
                    cell.setFlags(_NON_EDIT)
                    cell.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                    self.table.setItem(r_idx, c_idx, cell)

            elif spec[0] == "info":
                _, field_name, label = spec
                self.table.setRowHeight(r_idx, 30)
                lbl_item = QTableWidgetItem(f"  {label}")
                lbl_item.setBackground(QBrush(QColor(_INFO_BG)))
                lbl_item.setForeground(QBrush(QColor("#607080")))
                lbl_item.setFont(bold_font)
                lbl_item.setFlags(_NON_EDIT)
                self.table.setItem(r_idx, 0, lbl_item)
                for c_idx, vd in enumerate(vendors, start=1):
                    value = getattr(vd, field_name, "")
                    if isinstance(value, int):
                        display = f"{value} days" if value > 0 else "—"
                    else:
                        display = str(value) if value else "—"
                    cell = QTableWidgetItem(display)
                    cell.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                    cell.setBackground(QBrush(QColor(_INFO_BG)))
                    cell.setForeground(QBrush(QColor(_DARK)))
                    cell.setFlags(_NON_EDIT)
                    self.table.setItem(r_idx, c_idx, cell)

            elif spec[0] == "header":
                item = QTableWidgetItem(f"  {spec[1]}")
                item.setBackground(QBrush(QColor(_HDR_BG)))
                item.setForeground(QBrush(QColor(_HDR_FG)))
                item.setFont(hdr_font)
                item.setFlags(_NON_EDIT)
                self.table.setItem(r_idx, 0, item)
                self.table.setRowHeight(r_idx, 36)
                for c_idx in range(1, n_cols):
                    cell = QTableWidgetItem("")
                    cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    cell.setBackground(QBrush(QColor(_HDR_BG)))
                    cell.setForeground(QBrush(QColor(_DARK)))
                    cell.setFont(bold_font)
                    cell.setFlags(_NON_EDIT)
                    self.table.setItem(r_idx, c_idx, cell)

            elif spec[0] == "charge":
                _, bkt, name_lower, display = spec
                name_item = QTableWidgetItem(f"    {display}")
                name_item.setForeground(QBrush(QColor(_DARK)))
                name_item.setFlags(_NON_EDIT)
                self.table.setItem(r_idx, 0, name_item)
                self.table.setRowHeight(r_idx, 36)

                usd_vals: list[tuple[int, float]] = []

                for c_idx, vd in enumerate(vendors, start=1):
                    match = None
                    for ch in vd.charges:
                        if (ch.category == bkt and
                                ch.name_of_charge.strip().lower() == name_lower):
                            match = ch
                            break
                    if match and match.rate > 0 and not match.if_applicable:
                        usd_val = self._to_usd(match.rate, match.currency)
                        unit = match.unit_of_measurement

                        if _is_flat(unit):
                            flat_totals[c_idx] += usd_val
                        else:
                            vtot = variable_totals[c_idx]
                            vtot[unit] = vtot.get(unit, 0.0) + usd_val

                        cell = QTableWidgetItem(
                            _cell_display(match.rate, match.currency, unit)
                        )
                        cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        cell.setData(Qt.UserRole, (usd_val, unit))
                        cell.setFlags(_EDITABLE)

                        if match.remarks:
                            cell.setToolTip(f"Remarks: {match.remarks}")
                            cell.setBackground(QBrush(QColor(_YELLOW)))
                        else:
                            cell.setBackground(QBrush(QColor(_WHITE)))
                        usd_vals.append((c_idx, usd_val))
                    else:
                        cell = QTableWidgetItem("—")
                        cell.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                        cell.setForeground(QBrush(QColor(_DASH_FG)))
                        cell.setData(Qt.UserRole, None)
                        cell.setFlags(_EDITABLE)

                    self.table.setItem(r_idx, c_idx, cell)

                if len(usd_vals) >= 2:
                    min_val = min(v for _, v in usd_vals)
                    for c_idx, val in usd_vals:
                        if abs(val - min_val) < 0.001:
                            it = self.table.item(r_idx, c_idx)
                            if it and it.background().color().name() != _YELLOW.lower():
                                it.setBackground(QBrush(QColor(_GREEN)))
                                it.setForeground(QBrush(QColor(_GREEN_FG)))

            elif spec[0] == "total_usd":
                self._fill_total_row(r_idx, n_cols, flat_totals, variable_totals,
                                     "total_usd", bold_font)

            elif spec[0] == "total_inr":
                self._fill_total_row(r_idx, n_cols, flat_totals, variable_totals,
                                     "total_inr", bold_font)

        self.table.resizeColumnsToContents()
        self.table.setColumnWidth(0, max(240, self.table.columnWidth(0)))
        for c in range(1, n_cols):
            self.table.setColumnWidth(c, max(200, self.table.columnWidth(c)))

        self.table.blockSignals(False)
        self._update_totals()

    def _fill_total_row(self, r_idx: int, n_cols: int,
                        flat_totals: dict, variable_totals: dict,
                        spec_type: str, bold_font: QFont):
        inr_rate = self._get_usd_to_inr()
        mode = self._current_mode

        if spec_type == "total_usd":
            if mode == "air":
                wt = self._charge_weight
                lbl_txt = f"  Total USD  (at {wt:,.0f} KG)" if wt > 0 else "  Total USD"
            elif mode == "lcl":
                cbm = self._chargeable_cbm
                lbl_txt = f"  Total USD  (at {cbm:.3f} CBM)" if cbm > 0 else "  Total USD"
            else:
                lbl_txt = "  Total USD"
        else:
            if mode == "air":
                wt = self._charge_weight
                lbl_txt = f"  Total (Rs.)  (at {wt:,.0f} KG)" if wt > 0 else "  Total (Rs.)"
            elif mode == "lcl":
                cbm = self._chargeable_cbm
                lbl_txt = f"  Total (Rs.)  (at {cbm:.3f} CBM)" if cbm > 0 else "  Total (Rs.)"
            else:
                lbl_txt = "  Total (Rs.)"

        lbl = QTableWidgetItem(lbl_txt)
        lbl.setFont(bold_font)
        lbl.setBackground(QBrush(QColor(_TOT_BG)))
        lbl.setForeground(QBrush(QColor(_DARK)))
        lbl.setFlags(_NON_EDIT)
        self.table.setItem(r_idx, 0, lbl)

        needs_actual = self._needs_actual_total()
        row_h = 60 if needs_actual else 44
        self.table.setRowHeight(r_idx, row_h)

        pairs = []
        for c_idx in range(1, n_cols):
            flat = flat_totals.get(c_idx, 0.0)
            variable = variable_totals.get(c_idx, {})

            if spec_type == "total_usd":
                breakdown = _fmt_total_usd(flat, variable)
                if needs_actual:
                    actual = self._compute_actual(flat, variable)
                    display = f"{breakdown}\n= $ {actual:,.2f}"
                    cmp_val = actual
                else:
                    display = breakdown
                    cmp_val = flat + sum(variable.values())
            else:
                breakdown = _fmt_total_inr(flat, variable, inr_rate)
                if needs_actual:
                    actual_usd = self._compute_actual(flat, variable)
                    display = f"{breakdown}\n= Rs. {actual_usd * inr_rate:,.0f}"
                    cmp_val = actual_usd * inr_rate
                else:
                    display = breakdown
                    cmp_val = (flat + sum(variable.values())) * inr_rate

            cell = QTableWidgetItem(display)
            cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            cell.setFont(bold_font)
            cell.setBackground(QBrush(QColor(_TOT_BG)))
            cell.setForeground(QBrush(QColor(_DARK)))
            cell.setData(Qt.UserRole, cmp_val)
            cell.setFlags(_NON_EDIT)
            self.table.setItem(r_idx, c_idx, cell)
            pairs.append((c_idx, cmp_val))

        if len(pairs) >= 2:
            min_v = min(v for _, v in pairs)
            for c_idx, val in pairs:
                if abs(val - min_v) < 0.001 and val > 0:
                    it = self.table.item(r_idx, c_idx)
                    if it:
                        it.setBackground(QBrush(QColor(_GREEN)))
                        it.setForeground(QBrush(QColor(_GREEN_FG)))

    def _build_fcl_table(self, vendors: list, buckets: list):
        if not vendors:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            return

        # Group vendors by vendor_name preserving order
        vendor_groups: OrderedDict = OrderedDict()
        for vd in vendors:
            if vd.vendor_name not in vendor_groups:
                vendor_groups[vd.vendor_name] = []
            vendor_groups[vd.vendor_name].append(vd)

        ordered_vendors = [vd for group in vendor_groups.values() for vd in group]

        # Column headers = freight forwarder (vendor_name); shipping line shown as first data row
        col_labels = ["Charge"] + [vd.vendor_name for vd in ordered_vendors]
        n_cols = len(col_labels)

        # Collect charge rows across all buckets
        bucket_charge_rows: dict[str, list[tuple[str, str]]] = {b: [] for b in buckets}
        seen_per_bucket: dict[str, set] = defaultdict(set)

        for vd in ordered_vendors:
            for charge in vd.charges:
                if charge.if_applicable:
                    continue
                bkt = charge.category
                if bkt not in bucket_charge_rows:
                    continue
                name_lower = charge.name_of_charge.strip().lower()
                if name_lower and name_lower not in seen_per_bucket[bkt]:
                    seen_per_bucket[bkt].add(name_lower)
                    bucket_charge_rows[bkt].append(
                        (name_lower, charge.name_of_charge.strip())
                    )

        # Row specs: shipping_line_row first, then optional info rows, then buckets/charges
        row_specs: list = [("shipping_line_row",)]
        if any(vd.etd for vd in ordered_vendors):
            row_specs.append(("info", "etd", "ETD"))
        if any(vd.transit_days for vd in ordered_vendors):
            row_specs.append(("info", "transit_days", "Transit Time"))
        if any(vd.free_days_origin > 0 for vd in ordered_vendors):
            row_specs.append(("info", "free_days_origin", "Free Days (Origin)"))
        if any(vd.free_days_destination > 0 for vd in ordered_vendors):
            row_specs.append(("info", "free_days_destination", "Free Days (Destination)"))
        for bkt in buckets:
            if bucket_charge_rows[bkt]:
                row_specs.append(("header", bkt))
                for name_lower, display in bucket_charge_rows[bkt]:
                    row_specs.append(("charge", bkt, name_lower, display))
        row_specs.append(("total_usd",))
        row_specs.append(("total_inr",))
        self._row_specs = row_specs
        # Store vendor list for _update_totals and _add_vendor_column
        self._table_vendors = list(ordered_vendors)
        self._refresh_unit_inputs()

        flat_totals: dict[int, float] = {c: 0.0 for c in range(1, n_cols)}
        variable_totals: dict[int, dict[str, float]] = {c: {} for c in range(1, n_cols)}

        self.table.blockSignals(True)
        self.table.setColumnCount(n_cols)
        self.table.setHorizontalHeaderLabels(col_labels)
        self.table.setRowCount(len(row_specs))

        bold_font = QFont(); bold_font.setBold(True)
        hdr_font  = QFont(); hdr_font.setBold(True); hdr_font.setPointSize(10)

        for r_idx, spec in enumerate(row_specs):

            if spec[0] == "shipping_line_row":
                self.table.setRowHeight(r_idx, 28)
                lbl = QTableWidgetItem("  Shipping Line")
                lbl.setBackground(QBrush(QColor(_VHDR_BG)))
                lbl.setForeground(QBrush(QColor(_VHDR_FG)))
                lbl.setFont(hdr_font)
                lbl.setFlags(_NON_EDIT)
                self.table.setItem(r_idx, 0, lbl)
                for c_idx, vd in enumerate(ordered_vendors, start=1):
                    sl = vd.shipping_line or "—"
                    cell = QTableWidgetItem(sl)
                    cell.setBackground(QBrush(QColor(_VHDR_BG)))
                    cell.setForeground(QBrush(QColor(_VHDR_FG)))
                    cell.setFont(bold_font)
                    cell.setFlags(_NON_EDIT)
                    cell.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                    self.table.setItem(r_idx, c_idx, cell)

            elif spec[0] == "info":
                _, field_name, label = spec
                self.table.setRowHeight(r_idx, 30)
                lbl_item = QTableWidgetItem(f"  {label}")
                lbl_item.setBackground(QBrush(QColor(_INFO_BG)))
                lbl_item.setForeground(QBrush(QColor("#607080")))
                lbl_item.setFont(bold_font)
                lbl_item.setFlags(_NON_EDIT)
                self.table.setItem(r_idx, 0, lbl_item)
                for c_idx, vd in enumerate(ordered_vendors, start=1):
                    value = getattr(vd, field_name, "")
                    if isinstance(value, int):
                        display = f"{value} days" if value > 0 else "—"
                    else:
                        display = str(value) if value else "—"
                    cell = QTableWidgetItem(display)
                    cell.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                    cell.setBackground(QBrush(QColor(_INFO_BG)))
                    cell.setForeground(QBrush(QColor(_DARK)))
                    cell.setFlags(_NON_EDIT)
                    self.table.setItem(r_idx, c_idx, cell)

            elif spec[0] == "header":
                item = QTableWidgetItem(f"  {spec[1]}")
                item.setBackground(QBrush(QColor(_HDR_BG)))
                item.setForeground(QBrush(QColor(_HDR_FG)))
                item.setFont(hdr_font)
                item.setFlags(_NON_EDIT)
                self.table.setItem(r_idx, 0, item)
                self.table.setRowHeight(r_idx, 36)
                for c_idx in range(1, n_cols):
                    cell = QTableWidgetItem("")
                    cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    cell.setBackground(QBrush(QColor(_HDR_BG)))
                    cell.setForeground(QBrush(QColor(_DARK)))
                    cell.setFont(bold_font)
                    cell.setFlags(_NON_EDIT)
                    self.table.setItem(r_idx, c_idx, cell)

            elif spec[0] == "charge":
                _, bkt, name_lower, display = spec
                name_item = QTableWidgetItem(f"    {display}")
                name_item.setForeground(QBrush(QColor(_DARK)))
                name_item.setFlags(_NON_EDIT)
                self.table.setItem(r_idx, 0, name_item)
                self.table.setRowHeight(r_idx, 36)

                usd_vals: list[tuple[int, float]] = []

                for c_idx, vd in enumerate(ordered_vendors, start=1):
                    match = None
                    for ch in vd.charges:
                        if (ch.category == bkt and
                                ch.name_of_charge.strip().lower() == name_lower):
                            match = ch
                            break
                    if match and match.rate > 0 and not match.if_applicable:
                        usd_val = self._to_usd(match.rate, match.currency)
                        unit = match.unit_of_measurement

                        if _is_flat(unit):
                            flat_totals[c_idx] += usd_val
                        else:
                            vtot = variable_totals[c_idx]
                            vtot[unit] = vtot.get(unit, 0.0) + usd_val

                        cell = QTableWidgetItem(
                            _cell_display(match.rate, match.currency, unit)
                        )
                        cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        cell.setData(Qt.UserRole, (usd_val, unit))
                        cell.setFlags(_EDITABLE)

                        if match.remarks:
                            cell.setToolTip(f"Remarks: {match.remarks}")
                            cell.setBackground(QBrush(QColor(_YELLOW)))
                        else:
                            cell.setBackground(QBrush(QColor(_WHITE)))
                        usd_vals.append((c_idx, usd_val))
                    else:
                        cell = QTableWidgetItem("—")
                        cell.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                        cell.setForeground(QBrush(QColor(_DASH_FG)))
                        cell.setData(Qt.UserRole, None)
                        cell.setFlags(_EDITABLE)

                    self.table.setItem(r_idx, c_idx, cell)

                if len(usd_vals) >= 2:
                    min_val = min(v for _, v in usd_vals)
                    for c_idx, val in usd_vals:
                        if abs(val - min_val) < 0.001:
                            it = self.table.item(r_idx, c_idx)
                            if it and it.background().color().name() != _YELLOW.lower():
                                it.setBackground(QBrush(QColor(_GREEN)))
                                it.setForeground(QBrush(QColor(_GREEN_FG)))

            elif spec[0] == "total_usd":
                self._fill_total_row(r_idx, n_cols, flat_totals, variable_totals,
                                     "total_usd", bold_font)

            elif spec[0] == "total_inr":
                self._fill_total_row(r_idx, n_cols, flat_totals, variable_totals,
                                     "total_inr", bold_font)

        self.table.resizeColumnsToContents()
        self.table.setColumnWidth(0, max(240, self.table.columnWidth(0)))
        for c in range(1, n_cols):
            self.table.setColumnWidth(c, max(200, self.table.columnWidth(c)))

        self.table.blockSignals(False)
        self._update_totals()

    # ------------------------------------------------------------------
    # Cell editing
    # ------------------------------------------------------------------
    def _on_cell_changed(self, item: QTableWidgetItem):
        r, c = item.row(), item.column()
        if c == 0 or r >= len(self._row_specs):
            return
        spec = self._row_specs[r]
        if spec[0] != "charge":
            return

        text = item.text().strip()
        svc = self.app.currency_service

        self.table.blockSignals(True)
        if text and text not in ("—", "-"):
            parsed = _parse_cell_input(text, svc)
            if parsed:
                unit = parsed["unit"]
                item.setData(Qt.UserRole, (parsed["usd_val"], unit))
                item.setText(_cell_display(parsed["rate"], parsed["currency"], unit))
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item.setForeground(QBrush(QColor(_DARK)))
                if item.toolTip():
                    item.setToolTip("")
                    item.setBackground(QBrush(QColor(_WHITE)))
            else:
                item.setText("—")
                item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                item.setForeground(QBrush(QColor(_DASH_FG)))
                item.setData(Qt.UserRole, None)
        else:
            item.setText("—")
            item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            item.setForeground(QBrush(QColor(_DASH_FG)))
            item.setData(Qt.UserRole, None)
        self.table.blockSignals(False)

        self._update_totals()

    def _update_totals(self):
        n_cols = self.table.columnCount()
        flat_totals: dict[int, float] = defaultdict(float)
        variable_totals: dict[int, dict[str, float]] = defaultdict(dict)
        bucket_flat_totals: dict[str, dict[int, float]] = defaultdict(
            lambda: defaultdict(float)
        )
        bucket_variable_totals: dict[str, dict[int, dict[str, float]]] = defaultdict(
            lambda: defaultdict(dict)
        )

        for r_idx, spec in enumerate(self._row_specs):
            if spec[0] != "charge":
                continue
            bucket = spec[1]
            for c_idx in range(1, n_cols):
                it = self.table.item(r_idx, c_idx)
                if not it:
                    continue
                data = it.data(Qt.UserRole)
                if data is None:
                    continue
                usd_val, unit = data
                if _is_flat(unit):
                    flat_totals[c_idx] += usd_val
                    bucket_flat_totals[bucket][c_idx] += usd_val
                else:
                    variable_totals[c_idx][unit] = (
                        variable_totals[c_idx].get(unit, 0.0) + usd_val
                    )
                    bucket_vtot = bucket_variable_totals[bucket][c_idx]
                    bucket_vtot[unit] = bucket_vtot.get(unit, 0.0) + usd_val

        inr_rate = self._get_usd_to_inr()
        mode = self._current_mode
        needs_actual = self._needs_actual_total()

        bold_font = QFont(); bold_font.setBold(True)

        self.table.blockSignals(True)
        for r_idx, spec in enumerate(self._row_specs):
            if spec[0] == "vendor_header":
                continue

            elif spec[0] == "header":
                bucket = spec[1]
                lbl = self.table.item(r_idx, 0)
                if lbl:
                    lbl.setText(f"  {bucket}")
                self.table.setRowHeight(r_idx, 54 if needs_actual else 36)
                for c_idx in range(1, n_cols):
                    flat = bucket_flat_totals[bucket][c_idx]
                    variable = bucket_variable_totals[bucket][c_idx]
                    it = self.table.item(r_idx, c_idx)
                    if not it:
                        continue
                    if flat > 0 or variable:
                        breakdown = _fmt_total_usd(flat, variable)
                        if needs_actual:
                            actual = self._compute_actual(flat, variable)
                            display = f"{breakdown}\n= $ {actual:,.2f}"
                        else:
                            display = breakdown
                        it.setText(display)
                        it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        it.setFont(bold_font)
                        it.setBackground(QBrush(QColor(_HDR_BG)))
                        it.setForeground(QBrush(QColor(_DARK)))
                    else:
                        it.setText("—")
                        it.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                        it.setFont(bold_font)
                        it.setBackground(QBrush(QColor(_HDR_BG)))
                        it.setForeground(QBrush(QColor(_DASH_FG)))

            elif spec[0] == "total_usd":
                lbl = self.table.item(r_idx, 0)
                if lbl:
                    if mode == "air" and self._charge_weight > 0:
                        lbl.setText(f"  Total USD  (at {self._charge_weight:,.0f} KG)")
                    elif mode == "lcl" and self._chargeable_cbm > 0:
                        lbl.setText(f"  Total USD  (at {self._chargeable_cbm:.3f} CBM)")
                    else:
                        lbl.setText("  Total USD")
                self.table.setRowHeight(r_idx, 60 if needs_actual else 44)
                pairs = []
                for c_idx in range(1, n_cols):
                    flat = flat_totals[c_idx]
                    variable = variable_totals[c_idx]
                    breakdown = _fmt_total_usd(flat, variable)
                    if needs_actual:
                        actual = self._compute_actual(flat, variable)
                        display = f"{breakdown}\n= $ {actual:,.2f}"
                        cmp_val = actual
                    else:
                        display = breakdown
                        cmp_val = flat + sum(variable.values())
                    it = self.table.item(r_idx, c_idx)
                    if it:
                        it.setText(display)
                        it.setData(Qt.UserRole, cmp_val)
                        it.setBackground(QBrush(QColor(_TOT_BG)))
                        it.setForeground(QBrush(QColor(_DARK)))
                    pairs.append((c_idx, cmp_val))
                if len(pairs) >= 2:
                    min_v = min(v for _, v in pairs)
                    for c_idx, val in pairs:
                        if abs(val - min_v) < 0.001 and val > 0:
                            it = self.table.item(r_idx, c_idx)
                            if it:
                                it.setBackground(QBrush(QColor(_GREEN)))
                                it.setForeground(QBrush(QColor(_GREEN_FG)))

            elif spec[0] == "total_inr":
                lbl = self.table.item(r_idx, 0)
                if lbl:
                    if mode == "air" and self._charge_weight > 0:
                        lbl.setText(f"  Total (Rs.)  (at {self._charge_weight:,.0f} KG)")
                    elif mode == "lcl" and self._chargeable_cbm > 0:
                        lbl.setText(f"  Total (Rs.)  (at {self._chargeable_cbm:.3f} CBM)")
                    else:
                        lbl.setText("  Total (Rs.)")
                self.table.setRowHeight(r_idx, 60 if needs_actual else 44)
                pairs = []
                for c_idx in range(1, n_cols):
                    flat = flat_totals[c_idx]
                    variable = variable_totals[c_idx]
                    breakdown = _fmt_total_inr(flat, variable, inr_rate)
                    if needs_actual:
                        actual_usd = self._compute_actual(flat, variable)
                        display = f"{breakdown}\n= Rs. {actual_usd * inr_rate:,.0f}"
                        cmp_val = actual_usd * inr_rate
                    else:
                        display = breakdown
                        cmp_val = (flat + sum(variable.values())) * inr_rate
                    it = self.table.item(r_idx, c_idx)
                    if it:
                        it.setText(display)
                        it.setBackground(QBrush(QColor(_TOT_BG)))
                        it.setForeground(QBrush(QColor(_DARK)))
                    pairs.append((c_idx, cmp_val))
                if len(pairs) >= 2:
                    min_v = min(v for _, v in pairs)
                    for c_idx, val in pairs:
                        if abs(val - min_v) < 0.001 and val > 0:
                            it = self.table.item(r_idx, c_idx)
                            if it:
                                it.setBackground(QBrush(QColor(_GREEN)))
                                it.setForeground(QBrush(QColor(_GREEN_FG)))

        self._highlight_cheapest()
        self.table.blockSignals(False)

    def _highlight_cheapest(self):
        n_cols = self.table.columnCount()
        for r_idx, spec in enumerate(self._row_specs):
            if spec[0] != "charge":
                continue
            usd_vals: list[tuple[int, float]] = []
            for c_idx in range(1, n_cols):
                it = self.table.item(r_idx, c_idx)
                if not it:
                    continue
                data = it.data(Qt.UserRole)
                if data is not None:
                    usd_val, _ = data
                    usd_vals.append((c_idx, usd_val))
                    if it.background().color().name() != _YELLOW.lower():
                        it.setBackground(QBrush(QColor(_WHITE)))
                        it.setForeground(QBrush(QColor(_DARK)))
                else:
                    it.setBackground(QBrush(QColor(_WHITE)))
                    it.setForeground(QBrush(QColor(_DASH_FG)))

            if len(usd_vals) >= 2:
                min_v = min(v for _, v in usd_vals)
                for c_idx, val in usd_vals:
                    if abs(val - min_v) < 0.001:
                        it = self.table.item(r_idx, c_idx)
                        if it and it.background().color().name() != _YELLOW.lower():
                            it.setBackground(QBrush(QColor(_GREEN)))
                            it.setForeground(QBrush(QColor(_GREEN_FG)))

    # ------------------------------------------------------------------
    # Add / delete rows and columns
    # ------------------------------------------------------------------
    def _reset_spans(self):
        self.table.clearSpans()

    def _add_vendor_column(self):
        dlg = QInputDialog(self)
        dlg.setWindowTitle("Add Vendor")
        dlg.setLabelText("Vendor name:")
        dlg.setOkButtonText("Add")
        dlg.setTextValue("")
        dlg.setStyleSheet(_POPUP_STYLE)
        if dlg.exec() != QDialog.Accepted:
            return
        name = dlg.textValue().strip()
        if not name:
            return
        col = self.table.columnCount()

        bold_font = QFont(); bold_font.setBold(True)
        hdr_font  = QFont(); hdr_font.setBold(True); hdr_font.setPointSize(10)

        self.table.blockSignals(True)
        self.table.insertColumn(col)
        self.table.setHorizontalHeaderItem(col, QTableWidgetItem(name))

        for r_idx, spec in enumerate(self._row_specs):
            if spec[0] in ("vendor_header", "shipping_line_row", "airline_row"):
                cell = QTableWidgetItem("—")
                cell.setBackground(QBrush(QColor(_VHDR_BG)))
                cell.setForeground(QBrush(QColor(_VHDR_FG)))
                cell.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                cell.setFlags(_NON_EDIT)
                self.table.setItem(r_idx, col, cell)
            elif spec[0] == "info":
                cell = QTableWidgetItem("—")
                cell.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                cell.setBackground(QBrush(QColor(_INFO_BG)))
                cell.setForeground(QBrush(QColor(_DARK)))
                cell.setFlags(_NON_EDIT)
                self.table.setItem(r_idx, col, cell)
            elif spec[0] == "header":
                cell = QTableWidgetItem()
                cell.setBackground(QBrush(QColor(_HDR_BG)))
                cell.setFlags(_NON_EDIT)
                self.table.setItem(r_idx, col, cell)
            elif spec[0] == "charge":
                cell = QTableWidgetItem("—")
                cell.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                cell.setForeground(QBrush(QColor(_DASH_FG)))
                cell.setData(Qt.UserRole, None)
                cell.setFlags(_EDITABLE)
                self.table.setItem(r_idx, col, cell)
            elif spec[0] in ("total_usd", "total_inr"):
                txt = "$ 0.00" if spec[0] == "total_usd" else "Rs. 0"
                cell = QTableWidgetItem(txt)
                cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                cell.setFont(bold_font)
                cell.setBackground(QBrush(QColor(_TOT_BG)))
                cell.setForeground(QBrush(QColor(_DARK)))
                cell.setFlags(_NON_EDIT)
                self.table.setItem(r_idx, col, cell)

        self.table.blockSignals(False)
        self._reset_spans()
        self.table.setColumnWidth(col, 200)
        self._update_totals()

    def _add_charge_row(self):
        dlg = _AddChargeDialog(self._current_mode, self)
        if dlg.exec() != QDialog.Accepted:
            return
        bucket, charge_name = dlg.get_values()
        if not charge_name:
            return

        mode = self._current_mode
        buckets = FCL_BUCKETS if mode == "fcl" else (LCL_BUCKETS if mode == "lcl" else AIR_IMPORT_BUCKETS)

        # Find insert point: after last charge in this bucket
        insert_after = -1
        in_bucket = False
        for r_idx, spec in enumerate(self._row_specs):
            if spec[0] == "header" and spec[1] == bucket:
                insert_after = r_idx
                in_bucket = True
            elif in_bucket and spec[0] == "charge" and spec[1] == bucket:
                insert_after = r_idx
            elif in_bucket and spec[0] in ("header", "vendor_header", "shipping_line_row", "airline_row", "info", "total_usd", "total_inr"):
                break

        if insert_after < 0:
            tot_idx = next(
                (i for i, s in enumerate(self._row_specs) if s[0] == "total_usd"),
                len(self._row_specs) - 2,
            )
            self._row_specs.insert(tot_idx, ("header", bucket))
            self.table.blockSignals(True)
            self.table.insertRow(tot_idx)
            n_cols = self.table.columnCount()
            hdr_font = QFont(); hdr_font.setBold(True); hdr_font.setPointSize(10)
            hdr_item = QTableWidgetItem(f"  {bucket}")
            hdr_item.setBackground(QBrush(QColor(_HDR_BG)))
            hdr_item.setForeground(QBrush(QColor(_HDR_FG)))
            hdr_item.setFont(hdr_font)
            hdr_item.setFlags(_NON_EDIT)
            self.table.setItem(tot_idx, 0, hdr_item)
            self.table.setRowHeight(tot_idx, 30)
            for c_idx in range(1, n_cols):
                filler = QTableWidgetItem()
                filler.setBackground(QBrush(QColor(_HDR_BG)))
                filler.setFlags(_NON_EDIT)
                self.table.setItem(tot_idx, c_idx, filler)
            self.table.blockSignals(False)
            insert_after = tot_idx

        new_row_idx = insert_after + 1
        new_spec = ("charge", bucket, charge_name.lower(), charge_name)
        self._row_specs.insert(new_row_idx, new_spec)

        n_cols = self.table.columnCount()
        self.table.blockSignals(True)
        self.table.insertRow(new_row_idx)

        name_item = QTableWidgetItem(f"    {charge_name}")
        name_item.setForeground(QBrush(QColor(_DARK)))
        name_item.setFlags(_NON_EDIT)
        self.table.setItem(new_row_idx, 0, name_item)
        self.table.setRowHeight(new_row_idx, 36)

        for c_idx in range(1, n_cols):
            cell = QTableWidgetItem("—")
            cell.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            cell.setForeground(QBrush(QColor(_DASH_FG)))
            cell.setData(Qt.UserRole, None)
            cell.setFlags(_EDITABLE)
            self.table.setItem(new_row_idx, c_idx, cell)

        self.table.blockSignals(False)
        self._reset_spans()
        self._update_totals()

    def _delete_charge_row(self, r_idx: int):
        if r_idx < 0 or r_idx >= len(self._row_specs):
            return
        if self._row_specs[r_idx][0] != "charge":
            return
        self._row_specs.pop(r_idx)
        self.table.removeRow(r_idx)
        self._update_totals()

    def _delete_vendor_column(self, c_idx: int):
        if c_idx <= 0:
            return
        self.table.removeColumn(c_idx)
        self._update_totals()

    def _on_context_menu(self, pos):
        r_idx = self.table.rowAt(pos.y())
        c_idx = self.table.columnAt(pos.x())
        menu = QMenu(self)

        if 0 <= r_idx < len(self._row_specs):
            spec = self._row_specs[r_idx]
            if spec[0] == "charge":
                act = QAction(f"Delete row  '{spec[3]}'", self)
                act.triggered.connect(lambda: self._delete_charge_row(r_idx))
                menu.addAction(act)

        if c_idx > 0:
            hdr = self.table.horizontalHeaderItem(c_idx)
            vname = hdr.text() if hdr else ""
            act = QAction(f"Delete vendor  '{vname}'", self)
            act.triggered.connect(lambda: self._delete_vendor_column(c_idx))
            menu.addAction(act)

        if not menu.isEmpty():
            menu.exec(self.table.viewport().mapToGlobal(pos))

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------
    def _get_export_data(self) -> tuple[list, list[list]]:
        headers = []
        for c in range(self.table.columnCount()):
            h = self.table.horizontalHeaderItem(c)
            headers.append(h.text() if h else "")
        rows = []
        for r in range(self.table.rowCount()):
            row = []
            for c in range(self.table.columnCount()):
                item = self.table.item(r, c)
                row.append(item.text().strip() if item else "")
            rows.append(row)
        return headers, rows

    def _export_csv(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save CSV", "vendor_comparison.csv", "CSV Files (*.csv)"
        )
        if not path:
            return
        try:
            self._export_csv_to_path(path)
            self._show_message("Exported", f"Saved to:\n{path}")
        except Exception as exc:
            self._show_message("Export Error", str(exc), icon=QMessageBox.Critical)

    def _export_excel(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel", "vendor_comparison.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            self._export_excel_to_path(path)
            self._show_message("Exported", f"Saved to:\n{path}")
        except Exception as exc:
            self._show_message("Export Error", str(exc), icon=QMessageBox.Critical)

    def _show_message(self, title: str, text: str,
                      icon: QMessageBox.Icon = QMessageBox.Information):
        msg = QMessageBox(self)
        msg.setWindowTitle(title)
        msg.setText(text)
        msg.setIcon(icon)
        msg.setStyleSheet(_POPUP_STYLE)
        msg.exec()

    # ------------------------------------------------------------------
    # Non-interactive export helpers (used by UI wrappers and tests)
    # ------------------------------------------------------------------
    def _export_csv_to_path(self, path: str):
        headers, rows = self._get_export_data()
        # Build metadata rows: quote_type and shipping_line per vendor column
        quote_row = ["quote_type"]
        ship_row = ["shipping_line"]
        for c, h in enumerate(headers):
            if c == 0:
                continue
            vd = self._table_vendors[c - 1] if (hasattr(self, "_table_vendors") and len(self._table_vendors) >= c) else None
            if vd:
                quote_row.append(getattr(vd, "quote_type", ""))
                ship_row.append(getattr(vd, "shipping_line", ""))
            else:
                quote_row.append("")
                ship_row.append("")

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerow(quote_row)
            writer.writerow(ship_row)
            writer.writerows(rows)

    def _export_excel_to_path(self, path: str):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendor Comparison"

        headers, rows = self._get_export_data()
        n_cols = len(headers)

        thin = Side(style="thin", color="DCE3EA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        hdr_fill = PatternFill("solid", fgColor="1E2A3A")
        hdr_font = Font(color="FFFFFF", bold=True, size=11)
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        green_fill = PatternFill("solid", fgColor="C8E6C9")
        yellow_fill = PatternFill("solid", fgColor="FFF8E1")
        bkt_fill = PatternFill("solid", fgColor="E8EEF6")
        tot_fill = PatternFill("solid", fgColor="F0F4F8")
        vhdr_fill = PatternFill("solid", fgColor="2C3E50")

        for r_idx, (spec, row) in enumerate(zip(self._row_specs, rows), start=2):
            # Default write all values first
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="right" if c_idx > 1 else "left",
                    vertical="center", wrap_text=True
                )

            # Special styling / merging for FCL vendor header
            if spec[0] == "vendor_header" and getattr(self, "_current_mode", "") == "fcl" and hasattr(self, "_fcl_vendor_groups"):
                # Charge label cell
                c = ws.cell(row=r_idx, column=1)
                c.fill = vhdr_fill
                c.font = Font(bold=True, size=11, color="ECEFF1")
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = border

                c_offset = 2
                for vname, group in self._fcl_vendor_groups.items():
                    span = len(group)
                    # Set the left-most cell value for this vendor group
                    start_cell = ws.cell(row=r_idx, column=c_offset, value=vname)
                    start_cell.fill = vhdr_fill
                    start_cell.font = Font(bold=True, size=11, color="ECEFF1")
                    start_cell.alignment = Alignment(horizontal="center", vertical="center")
                    start_cell.border = border
                    if span > 1:
                        ws.merge_cells(start_row=r_idx, start_column=c_offset,
                                       end_row=r_idx, end_column=c_offset + span - 1)
                        # Apply border/fill to the merged range cells
                        for cc in range(c_offset, c_offset + span):
                            mc = ws.cell(row=r_idx, column=cc)
                            mc.fill = vhdr_fill
                            mc.font = Font(bold=True, size=11, color="ECEFF1")
                            mc.alignment = Alignment(horizontal="center", vertical="center")
                            mc.border = border
                    c_offset += span

            elif spec[0] == "vendor_header":
                # Fallback: style the whole row as vendor header
                for c_idx in range(1, n_cols + 1):
                    c = ws.cell(row=r_idx, column=c_idx)
                    c.fill = vhdr_fill
                    c.font = Font(bold=True, size=11, color="ECEFF1")
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border = border

            elif spec[0] == "header":
                for c_idx in range(1, n_cols + 1):
                    c = ws.cell(row=r_idx, column=c_idx)
                    c.fill = bkt_fill
                    c.alignment = Alignment(
                        horizontal="left" if c_idx == 1 else "right",
                        vertical="center",
                        wrap_text=True,
                    )
                    if c_idx == 1:
                        c.font = Font(bold=True, size=11, color="1565C0")
                    else:
                        c.font = Font(bold=True, size=11, color="1E2A3A")

            elif spec[0] in ("total_usd", "total_inr"):
                for c_idx in range(1, n_cols + 1):
                    c = ws.cell(row=r_idx, column=c_idx)
                    c.fill = tot_fill
                    c.font = Font(bold=True, size=11)

            # Apply green/yellow colouring based on UI background where present
            for c_idx in range(2, n_cols + 1):
                qt_item = self.table.item(r_idx - 2, c_idx - 1)
                if qt_item:
                    bg = qt_item.background().color().name()
                    if bg == _GREEN.lower():
                        ws.cell(row=r_idx, column=c_idx).fill = green_fill
                    elif bg == _YELLOW.lower():
                        ws.cell(row=r_idx, column=c_idx).fill = yellow_fill

        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) for c in col if c.value), default=10
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        ws.row_dimensions[1].height = 22

        wb.save(path)

    # ------------------------------------------------------------------
    @staticmethod
    def _pri_btn_style():
        return (
            "QPushButton{background:#1976D2;color:white;border:none;"
            "border-radius:6px;padding:0 18px;font-size:13px;font-weight:600;}"
            "QPushButton:hover{background:#1565C0;}"
        )

    @staticmethod
    def _sec_btn_style():
        return (
            "QPushButton{background:#FFFFFF;color:#1E2A3A;border:1px solid #90A4AE;"
            "border-radius:6px;padding:0 16px;font-size:13px;}"
            "QPushButton:hover{background:#F0F4F8;}"
        )

    @staticmethod
    def _sm_btn_style():
        return (
            "QPushButton{background:#EEF2F7;color:#1E2A3A;border:1px solid #B0BEC5;"
            "border-radius:5px;font-size:11px;padding:0 10px;}"
            "QPushButton:hover{background:#E3F2FD;}"
        )
