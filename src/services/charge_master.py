from pathlib import Path
import openpyxl


class ChargeMaster:
    """Loads the Freight_Forwarding_Charge_Master.xlsx and provides lookups."""

    def __init__(self, excel_path: str):
        self._path = excel_path
        self.buckets: list[str] = []
        # raw_lower -> {standardized, bucket}
        self._raw_map: dict[str, dict] = {}
        # bucket -> [standardized charge names] (ordered, deduplicated)
        self.bucket_charges: dict[str, list[str]] = {}
        self._load()

    def _load(self):
        wb = openpyxl.load_workbook(self._path, read_only=True, data_only=True)

        # 1. Ordered bucket list from Bucket Summary sheet
        ws_buckets = wb["Bucket Summary"]
        for row in ws_buckets.iter_rows(min_row=2, values_only=True):
            bucket = row[0]
            if bucket:
                self.buckets.append(str(bucket).strip())
                self.bucket_charges[str(bucket).strip()] = []

        # 2. Raw → standardized mapping from Charge Master sheet
        ws = wb["Charge Master"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw, standardized, bucket = row[0], row[1], row[2]
            if not (raw and standardized and bucket):
                continue
            raw_key = str(raw).lower().strip()
            std = str(standardized).strip()
            bkt = str(bucket).strip()
            self._raw_map[raw_key] = {"standardized": std, "bucket": bkt}
            if bkt in self.bucket_charges and std not in self.bucket_charges[bkt]:
                self.bucket_charges[bkt].append(std)

        wb.close()

    # ------------------------------------------------------------------
    def lookup(self, raw_charge: str) -> dict:
        """Return {standardized, bucket} or {} if not found."""
        return self._raw_map.get(raw_charge.lower().strip(), {})

    def get_charges_for_bucket(self, bucket: str) -> list[str]:
        return self.bucket_charges.get(bucket, [])

    def get_all_standardized(self) -> list[str]:
        seen, result = set(), []
        for charges in self.bucket_charges.values():
            for c in charges:
                if c not in seen:
                    seen.add(c)
                    result.append(c)
        return result

    def get_prompt_context(self) -> str:
        lines = ["CHARGE BUCKETS AND STANDARD CHARGE NAMES:"]
        for bucket in self.buckets:
            charges = self.bucket_charges[bucket]
            lines.append(f"\n{bucket}:")
            for c in charges:
                lines.append(f"  - {c}")
        return "\n".join(lines)
